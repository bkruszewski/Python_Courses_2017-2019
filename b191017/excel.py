import openpyxl

plik = openpyxl.load_workbook("example.xlsx")

arkusze = plik.get_sheet_names()

print(arkusze)

arkusz = plik.get_sheet_by_name("Owoce")

print(f"Aktywny arkusz: {arkusz.title}")

print(plik.active)

#komorki

komorka = arkusz["A1"]
print(komorka)
print(komorka.value)

#koordynaty

print(f"Adres komorki: {komorka.coordinate}")
print(f"Kolumna komorki: {komorka.column}")
print(f"Wiersz komorki: {komorka.row}")

#komorka

print (arkusz.cell(row=3, column=2))

#rozmiar arkusza

print(arkusz.max_column)
print(arkusz.max_row)

#zmiana liter na cyfry i vice-versa

from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(100))

print(column_index_from_string("AAG"))
print(column_index_from_string("AA"))

print(arkusz ["A1": "C2"]) # zwraca tupla z wierszami

for wiersz in arkusz["A1": "C7"]:
    for cell in wiersz:
        print(f"Komorka {cell.coordinate} ma wartosc {cell.value}")

arkusz["D3"] = "Moja wartosc"
print(arkusz["D3"].value)

plik.save("example2.xlsx")
